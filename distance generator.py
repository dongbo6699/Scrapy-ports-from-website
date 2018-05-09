import requests
import bs4
import openpyxl
import pickle

def get_url(url):
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/62.0.3202.94 Safari/537.36',
               'referer': 'http://ports.com/sea-route/',
               'Pragma': 'no-cache',
               'Host': 'ports.com',
               'X-Requested-With': 'XMLHttpRequest'}

    res = requests.get(url, headers=headers)
    
    return res


def get_num(city):
    headers = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/57.0.2987.98 Safari/537.36'}
    url='http://ports.com/aj/findport/?q='+city
    #print(url)
    res = requests.get(url, headers=headers)
    #print(res.text)
    #print(type(res.text))
    res=res.text

    res=res.splitlines()
    #print(res)
    i=0
    while i < len(res):
        if res[i].find('Port of '+city)!= -1:
            res=res[i].split(sep='|',maxsplit=2)
        elif res[i].find('Port '+city)!= -1:
            res=res[i].split(sep='|',maxsplit=2)
        elif res[i].find('Port of Goteborg')!= -1:
            res=res[i].split(sep='|',maxsplit=2)
        i+=1
   
    #print(res)
    '''
    if res.find(city)!= -1:
        num=res[res.find('Singapore'):33]
        '''
    num=res[1]
    #print(num)

    return num

def cal(city1,city2):
    num1 = get_num(city1)
    num2 = get_num(city2)
    host = 'http://ports.com/aj/sea-route/?'
    a= 'a=' + num1 + '&amp;'
    b= 'b=' + num2 + '&amp;'
    c= 'c=Port%20of%20' + city1 + ',%20' + '&amp;'
    d= 'd=Port%20of%20' + city2 + '&amp;'
    url = host+a+b+c+d
    #url = 'http://ports.com/aj/sea-route/?a=4595&amp;b=857&amp;c=Port%20of%20Shanghai,%20China&amp;d=Port%20of%20Singapore,%20Singapore'
    #'http://ports.com/aj/sea-route/?a=4595&amp;b=857&amp;c=Port%20of%20Shanghai,%20China&amp;d=Port%20of%20Singapore,%20Singapore'
    #'http://ports.com/aj/sea-route/?a=0&amp;b=0&amp;c=Port%20of%20New%20York&amp;d=Port%20of%20Shanghai'
    #input("请输入链接地址：")
    res = get_url(url)
    with open("res.txt", "w", encoding="utf-8") as file:
        file.write(res.text)    
    t=res.json()
    #print(type(t))

    for each in t.values():
        
        if each != "Could not find a route!":
            distance = t['cost']['nauticalmiles']
            twoport = t['title']
            #print('Distance from '+ twoport + ' is ' + distance + ' nauticalmiles')
        else:
            distance='no route'
    result=[]
    result.append([city1,city2,distance])
    return result
    
'''
    with open("distance.txt", "a", encoding="utf-8") as file:
        for each in result:
            file.write(each+'\n') 
'''
    


def save_to_excel(result):
    wb=openpyxl.Workbook()
    ws=wb.active
    ws['A1']='FromPort'
    ws['B1']='ToPort'
    ws['C1']='Distance'
    for each in result:
        ws.append(each)
    wb.save('MyDistance3.xlsx')

def pickle_data(result):
    pickle_file=open('C:\\Users\\bod\\Desktop\\instance generator\\distance.pkl','wb')
    pickle.dump(result,pickle_file)
    pickle_file.close()
    data2=[]
    f_b=open('C:\\Users\\bod\\Desktop\\instance generator\\distance.pkl','rb')
    data2 = pickle.load(f_b)  # 读取文件的二进制代码转换为相应数据类型
    print(data2)
    
def main():
    '''
    port=['Piraeus','Barcelona','Bremerhaven','Zeebrugge','Southampton',
          'Halifax','New York','Baltimore','Charleston','Brunswick','Miami','Manzanillo','Hueneme',
          'Yokohama','Kobe','Shanghai','Guangzhou','Laem Chabang','Singapore',
          'Auckland','Brisbane','Melbourne','Fremantle']
    '''
    port=['Piraeus','Barcelona','Bremerhaven','Gothenburg','Zeebrugge','Southampton',
          'Halifax','New York','Baltimore','Charleston','Brunswick','Miami','Manzanillo','Hueneme',
          'Yokohama','Kobe','Shanghai','Guangzhou','Laem Chabang','Singapore',
          'Auckland','Brisbane','Kembla','Melbourne','Fremantle']
          
    
    port=set(port)
    #print(port)
    result=[]
    for city1 in port:
        for city2 in port:
            #cal(city1,city2)
    
            result.extend(cal(city1,city2))
    #save_to_excel(result)
    #pickle_data(result)
    print(result)
    '''
    city1 = 'Piraeus'
    #country2 = 'China'
    
    city2 ='Singapore'
    #country2 ='Singapore'
    
    city2 ='Barcelona'
    #country1 ='Singapore'
    '''

    '''  
    with open("res.txt", "w", encoding="utf-8") as file:
        file.write(res.text)
    '''
    
    
if __name__ == "__main__":
    main()

